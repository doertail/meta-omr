import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

CIRCLED = {'①': '1', '②': '2', '③': '3', '④': '4', '⑤': '5'}


def _normalize(cell, allow_all=False):
    """셀 값을 정규화. 1~5 또는 allow_all=True일 때 모든 자연수 문자열 반환"""
    if cell is None:
        return None
    cell = str(cell).strip()
    if cell in CIRCLED:
        return CIRCLED[cell]
    if allow_all:
        if cell.isdecimal():
            return cell
    else:
        if cell in {'1', '2', '3', '4', '5'}:
            return cell
    return None


def _is_answer_table(table, allow_all=False):
    """테이블이 정답표인지 판별: 1~45 범위 정수 + 답이 충분히 있어야 함"""
    nums, answers = [], []
    for row in table:
        for cell in (row or []):
            s = str(cell or '').strip()
            if s.isdecimal() and 1 <= int(s) <= 45:
                nums.append(int(s))
            if _normalize(cell, allow_all):
                answers.append(cell)
    return len(nums) >= 5 and len(answers) >= 5


def _parse_answer_table(table, answer_map, allow_all=False):
    """정답표 파싱: 인터리브 포맷(1,②,2,⑤,...) 또는 행분리 포맷 처리"""
    nums_row = []
    for row in table:
        cells = [str(c or '').strip() for c in row]

        # 1단계: 인터리브 포맷 탐지 — 번호 바로 뒤에 원문자가 오는 쌍
        # 예: [1, ②, 2, ⑤, 3, ⑤, 4, ③, 5, ①]
        interleaved = []
        i = 0
        while i < len(cells) - 1:
            if cells[i].isdecimal() and 1 <= int(cells[i]) <= 45 and cells[i + 1] in CIRCLED:
                interleaved.append((int(cells[i]), CIRCLED[cells[i + 1]]))
                i += 2
            else:
                i += 1

        if interleaved:
            for n, a in interleaved:
                answer_map.setdefault(n, [])
                if a not in answer_map[n]:
                    answer_map[n].append(a)
            nums_row = []
            continue

        # 2단계: 행 분리 포맷 — 번호 행과 정답 행이 교대
        nums = [int(c) for c in cells if c.isdecimal() and 1 <= int(c) <= 45]
        ans = [_normalize(c, allow_all) for c in cells if _normalize(c, allow_all)]

        # 번호 행 판별: 중복 없이 오름차순 (정답 숫자 1~5를 번호로 오인 방지)
        is_q_row = bool(nums) and len(set(nums)) == len(nums) and nums == sorted(nums)

        if is_q_row:
            nums_row = nums
        elif ans and nums_row:
            for n, a in zip(nums_row, ans):
                answer_map.setdefault(n, [])
                if a not in answer_map[n]:
                    answer_map[n].append(a)
            nums_row = []


def _parse_text_fallback(text, answer_map, allow_all=False):
    """텍스트에서 '01. ③' 패턴으로 정답 파싱 (테이블 추출 실패 시 사용)"""
    # 1. 객관식 패턴: 숫자(1~45) + 마침표/괄호/공백 + 원문자
    matches = re.findall(r'\b(\d{1,2})[.)]\s*([①②③④⑤])', text)
    for num_str, circled in matches:
        n = int(num_str)
        if 1 <= n <= 45:
            a = CIRCLED[circled]
            answer_map.setdefault(n, [])
            if a not in answer_map[n]:
                answer_map[n].append(a)
    
    # 2. 주관식 패턴 (allow_all=True일 때): '단답형' 섹션 이후 '번호: 숫자' 형태 탐지
    if allow_all:
        subjective_matches = re.findall(r'(\d{1,2})\s*번?\s*답:\s*(\d+)', text)
        for num_str, ans_str in subjective_matches:
            n = int(num_str)
            if 1 <= n <= 45:
                answer_map.setdefault(n, [])
                if ans_str not in answer_map[n]:
                    answer_map[n].append(ans_str)


def extract_answers_with_pdfplumber(solution_path, is_math=False):
    """전 페이지 스캔으로 정답표 파싱.
    {번호(int): [정답후보(str), ...]} 반환.
    """
    try:
        answer_map = {}
        with pdfplumber.open(solution_path) as pdf:
            # 1단계: 테이블 추출 시도 (전 페이지)
            for page in pdf.pages:
                for table in page.extract_tables():
                    if _is_answer_table(table, allow_all=is_math):
                        _parse_answer_table(table, answer_map, allow_all=is_math)

            # 2단계: 테이블 추출 실패 시 텍스트 정규식 폴백
            if not answer_map:
                check_pages = [pdf.pages[0]]
                if len(pdf.pages) > 1:
                    check_pages.append(pdf.pages[-1])
                for page in check_pages:
                    text = page.extract_text() or ''
                    _parse_text_fallback(text, answer_map, allow_all=is_math)

        return answer_map or None
    except Exception as e:
        print(f"  ❌ pdfplumber 오류: {e}")
        return None


def save_results(output_path, df, mismatch_rows):
    """분류결과 시트(검증 컬럼 포함) 업데이트 + 정답검증 시트 재작성"""
    # 분류결과 시트: 검증 컬럼이 추가된 df로 교체
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='분류결과', index=False)

    # 정답검증 시트: 불일치/실패 행만
    wb = load_workbook(output_path)
    if '정답검증' in wb.sheetnames:
        del wb['정답검증']
    if mismatch_rows:
        ws = wb.create_sheet('정답검증')
        cols = ['파일명', '번호', '엑셀_정답', '추출_정답', '비고']
        ws.append(cols)
        for row in mismatch_rows:
            ws.append([row[c] for c in cols])
    wb.save(output_path)


def build_path_map(target_folder):
    """폴더 내 모든 PDF 파일의 파일명 → 전체 경로 매핑 테이블 생성"""
    path_map = {}
    for root, _, files in os.walk(target_folder):
        for f in files:
            if f.endswith('.pdf'):
                path_map[f] = os.path.join(root, f)
    return path_map


def main():
    target_folder = "./고등 모의고사 기출"

    subject = input("검증할 과목을 입력하세요 (예: 국어, 수학, 영어): ").strip()
    if not subject:
        print("과목이 입력되지 않았습니다.")
        return

    output_path = f"./분류결과/분류결과_{subject}.xlsx"
    if not os.path.exists(output_path):
        print(f"❌ {output_path} 파일이 없습니다. class.py를 먼저 실행해주세요.")
        return

    try:
        df = pd.read_excel(output_path, sheet_name='분류결과')
    except Exception:
        df = pd.read_excel(output_path)

    if '정답' not in df.columns:
        print("❌ '정답' 컬럼이 없습니다. 최신 버전의 class.py로 분류결과를 다시 생성해주세요.")
        return

    print(f"✓ {len(df)}개 항목 로드 완료\n")

    print("📂 PDF 파일 경로 매핑 중...")
    path_map = build_path_map(target_folder)

    problem_files = df['파일명'].unique()
    print(f"✓ 검증 대상: {len(problem_files)}개 파일\n")

    mismatch_rows = []
    total_compared = 0
    total_mismatch = 0
    total_skipped = 0
    total_failed = 0

    # 검증 결과 컬럼 초기화
    df['검증_추출정답'] = ''
    df['검증_비고'] = ''

    for i, problem_filename in enumerate(problem_files, 1):
        solution_filename = problem_filename.replace('문제', '해설')
        solution_path = path_map.get(solution_filename)

        print(f"[{i}/{len(problem_files)}] {problem_filename}")

        file_rows = df[df['파일명'] == problem_filename]

        if not solution_path:
            df.loc[file_rows.index, '검증_비고'] = '해설없음'
            print(f"  ⚠ 해설 파일 없음 → 스킵\n")
            continue

        answer_map = extract_answers_with_pdfplumber(solution_path)

        if answer_map is None:
            df.loc[file_rows.index, '검증_비고'] = '추출실패'
            for _, row in file_rows.iterrows():
                mismatch_rows.append({
                    '파일명': problem_filename,
                    '번호': row['번호'],
                    '엑셀_정답': str(row['정답']),
                    '추출_정답': '',
                    '비고': '추출실패'
                })
            total_failed += len(file_rows)
            print(f"  ❌ 정답 추출 실패\n")
            continue

        file_mismatch = 0
        file_skipped = 0
        for idx, row in file_rows.iterrows():
            question_num = int(row['번호'])
            excel_answer = str(row['정답'])
            candidates = answer_map.get(question_num)

            if candidates is None:
                # 추출 안 됨 (선택과목 등) → 스킵
                df.loc[idx, '검증_비고'] = '스킵'
                file_skipped += 1
                total_skipped += 1
            elif excel_answer in candidates:
                # 일치
                df.loc[idx, '검증_추출정답'] = ' / '.join(candidates)
                df.loc[idx, '검증_비고'] = '일치'
                total_compared += 1
            else:
                # 불일치
                df.loc[idx, '검증_추출정답'] = ' / '.join(candidates)
                df.loc[idx, '검증_비고'] = '불일치'
                total_compared += 1
                total_mismatch += 1
                file_mismatch += 1
                mismatch_rows.append({
                    '파일명': problem_filename,
                    '번호': question_num,
                    '엑셀_정답': excel_answer,
                    '추출_정답': ' / '.join(candidates),
                    '비고': '불일치'
                })

        if file_mismatch == 0:
            print(f"  ✅ 전체 일치 ({len(file_rows) - file_skipped}문항 검증, {file_skipped}문항 스킵)")
        else:
            print(f"  ⚠ 불일치 {file_mismatch}건 발견 ({file_skipped}문항 스킵)")

        print()

    # 최종 저장: 분류결과(검증 컬럼 포함) + 정답검증 시트
    save_results(output_path, df, mismatch_rows)
    if mismatch_rows:
        print(f"✓ '정답검증' 시트 저장 완료: {len(mismatch_rows)}건\n")
    else:
        print("✅ 불일치 항목 없음.\n")

    # 통계
    print("=" * 60)
    print("📈 검증 통계")
    print("=" * 60)
    print(f"총 비교 항목: {total_compared}건")
    print(f"불일치: {total_mismatch}건")
    print(f"추출 스킵 (선택과목 등): {total_skipped}건")
    print(f"추출 실패: {total_failed}건")
    if total_compared > 0:
        accuracy = (total_compared - total_mismatch) / total_compared * 100
        print(f"일치율: {accuracy:.1f}%")


if __name__ == "__main__":
    main()
